"""
VT Schedules export — generate a schedule-style workbook listing every VT parameter
flagged with **column E = "yes"** in ``VT standard configurations_V00.xlsx``.

Layout of the produced ``.xlsx``:

- Row 1: ``Parameter | Unit | Value`` header (bold).
- One block per lift. Blocks are separated by a **bold** ``Shaft 1`` / ``Shaft 2`` … row
  in column A (same visual role as the LD export's shaft markers).
- Bold section-header rows — imported directly from VT column A rows that carry a label
  but no unit / export flag / LD path — preserve the VT document's grouping
  (General Specification, Electrical & HVAC, Mechanical Forces, etc.).
- Parameter rows:
  - column A = parameter name (VT column A)
  - column B = unit (VT column D, brackets stripped — e.g. ``[mm]`` → ``mm``)
  - column C = resolved value (same resolver used by the LD export, so UI inputs win
    over formula-derived fallbacks and over legacy lift-dict lookups)

Values are computed via :func:`lift_designer_ld_export._value_for_param` over the same
:class:`_ExportCtx` the LD export builds — guaranteeing the two exports stay in sync
for any parameter shared between them.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Mapping, Optional, Sequence, Set, Tuple


# Yellow/amber fill used to highlight cells whose value came from a non-standard
# dropdown override. Matches the UI highlight on :class:`OverrideComboBox`.
OVERRIDE_HIGHLIGHT_HEX: str = "FFFFF2CC"  # ARGB: FF + FFF2CC


def _normalize_override_label(label: Any) -> str:
    """Normalization shared by the UI and writer so overrides can be matched."""
    if label is None:
        return ""
    return " ".join(str(label).split()).lower()

from lift_designer_ld_export import (
    _ExportCtx,
    _coerce_cell_value,
    _compliance,
    _cost,
    _drive,
    _emergency,
    _forces,
    _lift,
    _value_for_param,
    default_vt_workbook_path,
)
from lift_designer_vt_derived import compute_derived


__all__ = [
    "VTScheduleRule",
    "ScheduleRow",
    "load_vt_schedule_rules",
    "build_schedule_rows_from_user_inputs",
    "build_schedule_rows_per_lift",
    "matrix_for_schedule_workbook",
    "matrix_for_schedule_workbook_multi",
    "write_schedule_workbook_multi",
    "default_schedule_filename",
    "default_schedule_template_path",
    "write_schedule_workbook_from_template",
    "TEMPLATE_SHEET_NAME",
    "TEMPLATE_LIFT_HEADER_ROW",
    "TEMPLATE_FIRST_LIFT_COL",
    "TEMPLATE_LIFT_COL_STRIDE",
    "TEMPLATE_MAX_LIFTS",
    "TEMPLATE_REVISION_FIRST_ROW",
    "TEMPLATE_REVISION_MAX_ROWS",
    "TEMPLATE_REVISION_COLUMNS",
    "REVISION_FIELDS",
    "suggest_next_revision_code",
    "OVERRIDE_HIGHLIGHT_HEX",
    "_normalize_override_label",
]


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class VTScheduleRule:
    """One row pulled from the VT sheet for the schedules export.

    ``kind`` is either ``"data"`` (parameter row with column E = ``yes``) or
    ``"section"`` (organizational header row).
    """

    kind: str
    param_label: str = ""
    unit: str = ""
    vt_row: int = 0


@dataclass(frozen=True)
class ScheduleRow:
    """Resolved schedule row ready for the output workbook."""

    param_label: str
    unit: str = ""
    value: str = ""
    vt_row: int = 0
    is_section: bool = False


# ---------------------------------------------------------------------------
# VT sheet scanning
# ---------------------------------------------------------------------------

def _strip_unit_brackets(raw: Any) -> str:
    """Return the unit text without wrapping brackets (``"[mm]"`` → ``"mm"``)."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s or s == "-":
        return ""
    m = re.match(r"^\[(.*)\]$", s)
    if m:
        return m.group(1).strip()
    return s


def _looks_like_section_header(a_s: str, d_s: str, e_s: str, f_s: str) -> bool:
    """
    Heuristic for a *pure* organizational row in the VT sheet.

    Section-header rows carry a label in column **A** but leave columns **D** (unit),
    **E** (schedule flag), and **F** (LD path) completely blank. Parameter rows that
    are simply excluded from export (``F = "no" / "n" / "-" / "?"`` or similar) still
    carry a value in column F and therefore must *not* be promoted to section headers
    — otherwise rows such as *Configuration short name* (``F = "No"``) would acquire a
    bold title style they don't deserve.
    """
    if not a_s:
        return False
    if d_s or e_s or f_s:
        return False
    return True


def _norm(s: str) -> str:
    return " ".join(s.lower().split())


# Section-header titles to drop from the schedules output entirely (case- and whitespace-
# insensitive match against VT column A). These rows look like section headers by the
# structural rule above but don't add value to the schedule document.
_EXCLUDED_SECTION_TITLES = {
    _norm(s) for s in (
        "MUZ???",
        "Mechanical Forces",
        "Interior Design Lifts",
        "Interior Design Escalator/Moving Walk",
        "Functions",
        "Lift Designer parameters",
        "wide cabin front panels:",
        "Handrail",
        "Schematics and Occupancies",
    )
}

# Rows that *structurally* look like section headers (D/E/F all blank) but are really
# single data rows that belong in the schedule without bold styling. They are emitted
# as ``data`` rules with empty unit; the value resolves through the standard pipeline.
_FORCE_DATA_ROWS = {
    _norm(s) for s in (
        "EN81-58 fire rating class",
    )
}


def load_vt_schedule_rules(path_vt: Optional[str] = None) -> List[VTScheduleRule]:
    """
    Scan ``VT Standard configs`` and return an ordered list of rules:

    - Every row where column **E** is ``yes`` becomes a ``data`` rule carrying the
      parameter label (column A), unit (column D, brackets stripped), and VT row.
    - Rows that look like section headers (see :func:`_looks_like_section_header`)
      become ``section`` rules so the output workbook can emit bold group titles in
      the same order they appear in VT.
    """
    path_vt = path_vt or default_vt_workbook_path()
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e

    wb = load_workbook(path_vt, data_only=True)
    if "VT Standard configs" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["VT Standard configs"]

    rules: List[VTScheduleRule] = []
    for r in range(3, ws.max_row + 1):
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        e = ws.cell(r, 5).value
        f = ws.cell(r, 6).value
        a_s = "" if a is None else str(a).strip()
        if not a_s:
            continue
        d_s = "" if d is None else str(d).strip()
        e_s = "" if e is None else str(e).strip()
        f_s = "" if f is None else str(f).strip()

        label_key = _norm(a_s)

        if e_s.lower() == "yes":
            rules.append(
                VTScheduleRule(
                    kind="data",
                    param_label=a_s,
                    unit=_strip_unit_brackets(d_s),
                    vt_row=r,
                )
            )
            continue

        if _looks_like_section_header(a_s, d_s, e_s, f_s):
            # User-curated overrides for rows that structurally look like section
            # headers but shouldn't be rendered as bold titles.
            if label_key in _EXCLUDED_SECTION_TITLES:
                continue
            if label_key in _FORCE_DATA_ROWS:
                rules.append(
                    VTScheduleRule(
                        kind="data",
                        param_label=a_s,
                        unit=_strip_unit_brackets(d_s),
                        vt_row=r,
                    )
                )
                continue
            rules.append(VTScheduleRule(kind="section", param_label=a_s, vt_row=r))

    wb.close()
    return rules


# ---------------------------------------------------------------------------
# Row building — reuses the LD export's value resolver
# ---------------------------------------------------------------------------

def build_schedule_rows_from_user_inputs(
    user_inputs: Mapping[str, Any],
    lift_index: int = 0,
    vt_path: Optional[str] = None,
    door_manufacturer: Optional[str] = None,
) -> List[ScheduleRow]:
    """
    Build the ordered schedule rows for one lift by resolving every ``yes`` parameter
    label through the LD export's :func:`_value_for_param`. Section rules pass through
    as section marker rows (``is_section=True``).

    ``door_manufacturer`` selects the VT R277 manufacturer cell. See
    :func:`lift_designer_vt_derived.compute_derived` for the resolution rules.
    """
    ctx = _ExportCtx(
        user_inputs=user_inputs,
        lift_index=lift_index,
        lift=_lift(user_inputs, lift_index),
        forces=_forces(user_inputs, lift_index),
        drive=_drive(user_inputs, lift_index),
        compliance=_compliance(user_inputs, lift_index),
        emergency=_emergency(user_inputs, lift_index),
        cost=_cost(user_inputs, lift_index),
        derived=compute_derived(user_inputs, lift_index, door_manufacturer=door_manufacturer),
    )
    rules = load_vt_schedule_rules(vt_path)
    rows: List[ScheduleRow] = []
    for rule in rules:
        if rule.kind == "section":
            rows.append(
                ScheduleRow(
                    param_label=rule.param_label,
                    vt_row=rule.vt_row,
                    is_section=True,
                )
            )
            continue
        val = _value_for_param(rule.param_label, ctx)
        rows.append(
            ScheduleRow(
                param_label=rule.param_label,
                unit=rule.unit,
                value=val,
                vt_row=rule.vt_row,
            )
        )
    return rows


def build_schedule_rows_per_lift(
    user_inputs: Mapping[str, Any],
    num_lifts: int,
    vt_path: Optional[str] = None,
    door_manufacturer: Optional[str] = None,
) -> List[List[ScheduleRow]]:
    """Return ``num_lifts`` row lists — one per lift index."""
    n = max(0, int(num_lifts))
    return [
        build_schedule_rows_from_user_inputs(
            user_inputs,
            lift_index=i,
            vt_path=vt_path,
            door_manufacturer=door_manufacturer,
        )
        for i in range(n)
    ]


# ---------------------------------------------------------------------------
# Matrix → workbook
# ---------------------------------------------------------------------------

_SCHEDULE_HEADERS: Tuple[str, str, str] = ("Parameter", "Unit", "Value")


def _row_section(title: str) -> List[Any]:
    """Empty row with the section title in column A; value/unit columns are blank."""
    return [title, None, None]


def _prune_trailing_sections(matrix: List[List[Any]], bold_rows: List[int]) -> None:
    """
    Drop any tail of section-only rows (section header immediately followed by another
    section header / end of block) so the workbook never ends with orphan bold titles.
    Runs in place on ``matrix`` and ``bold_rows``.
    """
    # Scan from the end: if the last non-preamble row is a section header (bold row
    # with empty B and C) whose next item is another section header or EOF, drop it.
    changed = True
    while changed:
        changed = False
        i = len(matrix) - 1
        while i >= 1:
            row = matrix[i]
            is_bold = (i + 1) in bold_rows
            is_pure_label = is_bold and (row[1] is None) and (row[2] is None)
            if not is_pure_label:
                break
            # Look at next row (below). If there is no next data row before another
            # bold row or EOF, this header is orphaned.
            has_child = False
            for j in range(i + 1, len(matrix)):
                nxt = matrix[j]
                if (j + 1) in bold_rows:
                    break
                if any(c not in (None, "") for c in nxt):
                    has_child = True
                    break
            if has_child:
                i -= 1
                continue
            # Drop this row
            matrix.pop(i)
            bold_rows[:] = [b if b < i + 1 else b - 1 for b in bold_rows if b != i + 1]
            changed = True
            break


def matrix_for_schedule_workbook(
    rows: Sequence[ScheduleRow],
) -> Tuple[List[List[Any]], List[int]]:
    """
    Build a 3-column grid for one lift block. Returns ``(matrix, bold_row_indices)``
    where ``bold_row_indices`` are 1-based Excel rows to render in bold (header row
    and section-title rows).

    Section headers that have no data row following them inside this block are
    pruned so the workbook doesn't carry empty bold titles.
    """
    matrix: List[List[Any]] = [list(_SCHEDULE_HEADERS)]
    bold_rows: List[int] = [1]  # header row

    for r in rows:
        if r.is_section:
            matrix.append(_row_section(r.param_label))
            bold_rows.append(len(matrix))
            continue
        matrix.append([r.param_label, r.unit or None, r.value or None])

    _prune_trailing_sections(matrix, bold_rows)
    return matrix, bold_rows


def matrix_for_schedule_workbook_multi(
    rows_by_lift: Sequence[Sequence[ScheduleRow]],
) -> Tuple[List[List[Any]], List[int]]:
    """
    Single sheet for several lifts. Preamble once (header row), then each lift's block
    prefixed with a bold ``Shaft N`` row in column A (same role as the LD export's
    shaft separators).
    """
    matrix: List[List[Any]] = [list(_SCHEDULE_HEADERS)]
    bold_rows: List[int] = [1]

    for i, rows in enumerate(rows_by_lift):
        matrix.append(_row_section(f"Shaft {i + 1}"))
        bold_rows.append(len(matrix))
        sub_m, sub_bold = matrix_for_schedule_workbook(list(rows))
        body = sub_m[1:]  # drop sub-preamble
        offset = len(matrix)
        matrix.extend(body)
        for b in sub_bold:
            if b > 1:
                bold_rows.append(offset + (b - 1))

    return matrix, bold_rows


def write_schedule_workbook_multi(
    path: str,
    rows_by_lift: Sequence[Sequence[ScheduleRow]],
    sheet_title: str = "VT Schedules",
) -> None:
    """
    Write the combined multi-lift schedule workbook to ``path``. Uses ``openpyxl``;
    columns are auto-sized and section / header rows are bolded.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e

    matrix, bold_rows = matrix_for_schedule_workbook_multi(rows_by_lift)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    font_bold = Font(bold=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    bold_set = set(bold_rows)

    for r_idx, line in enumerate(matrix, start=1):
        is_bold = r_idx in bold_set
        for c_idx, cell in enumerate(line, start=1):
            if c_idx > 3:
                continue
            cl = ws.cell(row=r_idx, column=c_idx, value=cell)
            cl.alignment = align_center if c_idx == 2 else align_left
            if is_bold and cell is not None:
                cl.font = font_bold

    for col in range(1, 4):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2.5, 12), 85)

    wb.save(path)


# ---------------------------------------------------------------------------
# File-name helper used by the UI
# ---------------------------------------------------------------------------

def default_schedule_filename(payload: Mapping[str, Any]) -> str:
    """Suggest a default file name ``<ProjectName>_Schedules.xlsx`` for the save dialog."""
    fn = payload.get("FileName", "project") if isinstance(payload, Mapping) else "project"
    base = "VT_schedules"
    if isinstance(fn, str) and fn.strip():
        base = os.path.splitext(fn.strip())[0] or base
    return f"{base}_Schedules.xlsx"


# ---------------------------------------------------------------------------
# Template-based export
#
# The project ships with a VT Schedules template workbook that defines the
# canonical layout for the schedules export (column A = parameter name in
# English, column D = German translation, column F = unit, row 8 = lift
# number header). The export fills one column per lift with values resolved
# through the same :func:`_value_for_param` pipeline the LD export uses, so
# UI inputs win over formula-derived fallbacks.
# ---------------------------------------------------------------------------

TEMPLATE_SHEET_NAME: str = "VT General Schedules"
TEMPLATE_LIFT_HEADER_ROW: int = 8           # Row whose column A is "Lift Number".
TEMPLATE_PARAM_FIRST_ROW: int = 9           # First parameter row below the header.
TEMPLATE_FIRST_LIFT_COL: int = 8            # Column H → lift 1.
TEMPLATE_LIFT_COL_STRIDE: int = 4           # H, L, P, T, X, AB, AF, AJ.
TEMPLATE_MAX_LIFTS: int = 8                 # Template has 8 pre-laid-out lift columns.

# --- Revision title block layout -----------------------------------------
# Row 1 is the header (Revision / Issue Purpose / Date / Design Eng. / Checked);
# rows 2..6 hold one revision each (row 7 is a spacer, row 8 is already the
# "Lift Number" header). Each field spans a merged 3-cell block in its row —
# writes only need to target the top-left cell of each merged range.
TEMPLATE_REVISION_FIRST_ROW: int = 2
TEMPLATE_REVISION_MAX_ROWS: int = 5         # Rows 2..6 inclusive.
TEMPLATE_REVISION_COLUMNS: Dict[str, int] = {
    "revision": 12,        # L
    "issue_purpose": 16,   # P
    "date": 20,            # T
    "design_eng": 24,      # X
    "checked": 28,         # AB
}
REVISION_FIELDS: Tuple[str, ...] = (
    "revision",
    "issue_purpose",
    "date",
    "design_eng",
    "checked",
)

# Candidate template filenames, searched in the repo root in the order listed.
# The first name is the canonical one for future-proofing; the second matches
# the file as currently uploaded to the project.
_TEMPLATE_FILENAMES: Tuple[str, ...] = (
    "VT Schedules template.xlsx",
    "xxxxxxx_template_VT Schedules V3.0_de_en.xlsx",
)


def default_schedule_template_path() -> str:
    """
    Return the absolute path to the VT Schedules template workbook that ships
    next to this module. Raises :class:`FileNotFoundError` if none of the known
    template filenames can be located.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for name in _TEMPLATE_FILENAMES:
        p = os.path.join(base_dir, name)
        if os.path.isfile(p):
            return p
    # Last-resort: any ``*template*VT*Schedules*.xlsx`` in the same folder.
    try:
        for fn in os.listdir(base_dir):
            lo = fn.lower()
            if (
                lo.endswith(".xlsx")
                and "template" in lo
                and "vt" in lo
                and "schedule" in lo
            ):
                return os.path.join(base_dir, fn)
    except OSError:
        pass
    raise FileNotFoundError(
        "VT Schedules template workbook not found. Expected one of: "
        + ", ".join(_TEMPLATE_FILENAMES)
        + f"  (looked in: {base_dir})"
    )


def _lift_column_index(lift_number: int) -> int:
    """
    Return the 1-based Excel column for the given 1-based lift number in the
    template layout (lift 1 → H, lift 2 → L, …).
    """
    if lift_number < 1:
        raise ValueError("lift_number is 1-based and must be >= 1")
    return TEMPLATE_FIRST_LIFT_COL + TEMPLATE_LIFT_COL_STRIDE * (lift_number - 1)


def suggest_next_revision_code(existing: Sequence[Mapping[str, Any]]) -> str:
    """
    Given the current list of stored schedule revisions, suggest the next code.

    If the most recent entry uses the ``Pnn`` pattern (``P01`` … ``P99``), the
    next integer is returned zero-padded. Otherwise we fall back to ``P<count+1>``
    with a leading zero so the very first export defaults to ``P01``.
    """
    n = len(existing)
    last = str(existing[-1].get("revision", "")).strip() if n else ""
    m = re.match(r"^[Pp](\d{1,3})$", last)
    if m:
        return f"P{int(m.group(1)) + 1:02d}"
    return f"P{n + 1:02d}"


def _clear_revision_block(ws) -> None:
    """
    Wipe every data cell (rows ``TEMPLATE_REVISION_FIRST_ROW`` …
    ``TEMPLATE_REVISION_FIRST_ROW + TEMPLATE_REVISION_MAX_ROWS - 1``) across the
    five revision columns so no template placeholder survives under a fresh list.
    """
    first = TEMPLATE_REVISION_FIRST_ROW
    last = first + TEMPLATE_REVISION_MAX_ROWS - 1
    for r in range(first, last + 1):
        for col in TEMPLATE_REVISION_COLUMNS.values():
            ws.cell(r, col).value = None


def _write_revision_block(ws, revisions: Sequence[Mapping[str, Any]]) -> int:
    """
    Write up to :data:`TEMPLATE_REVISION_MAX_ROWS` revisions into the title
    block (row 2 = first / oldest, next row = next revision, …). Returns the
    number of revisions actually written. Any extras beyond the template's
    capacity are silently dropped from the *front* of the list so the most
    recent entries always make it into the workbook.
    """
    if not revisions:
        return 0

    trimmed = list(revisions)[-TEMPLATE_REVISION_MAX_ROWS:]
    _clear_revision_block(ws)

    first = TEMPLATE_REVISION_FIRST_ROW
    cols = TEMPLATE_REVISION_COLUMNS
    for idx, rev in enumerate(trimmed):
        row = first + idx
        for key, col in cols.items():
            val = rev.get(key, "")
            if val in (None, ""):
                continue
            ws.cell(row, col).value = _coerce_cell_value(val)
    return len(trimmed)


def _build_ctx(
    user_inputs: Mapping[str, Any],
    lift_index: int,
    door_manufacturer: Optional[str] = None,
) -> _ExportCtx:
    """Assemble the per-lift resolver context used by :func:`_value_for_param`."""
    return _ExportCtx(
        user_inputs=user_inputs,
        lift_index=lift_index,
        lift=_lift(user_inputs, lift_index),
        forces=_forces(user_inputs, lift_index),
        drive=_drive(user_inputs, lift_index),
        compliance=_compliance(user_inputs, lift_index),
        emergency=_emergency(user_inputs, lift_index),
        cost=_cost(user_inputs, lift_index),
        derived=compute_derived(user_inputs, lift_index, door_manufacturer=door_manufacturer),
    )


def write_schedule_workbook_from_template(
    output_path: str,
    user_inputs: Mapping[str, Any],
    num_lifts: int,
    template_path: Optional[str] = None,
    revisions: Optional[Sequence[Mapping[str, Any]]] = None,
    overrides: Optional[Mapping[int, Sequence[str]]] = None,
    door_manufacturer: Optional[str] = None,
) -> int:
    """
    Fill the VT Schedules template with values resolved from ``user_inputs`` and
    save to ``output_path``. Returns the number of lifts actually written.

    Behaviour:

    - Loads ``template_path`` (defaults to :func:`default_schedule_template_path`)
      so the template's styling, merged cells, German translations and units are
      preserved.
    - For each 1-based lift ``N`` (up to :data:`TEMPLATE_MAX_LIFTS`), walks
      column A from row :data:`TEMPLATE_PARAM_FIRST_ROW` to the sheet's last row
      and, for every non-empty label, writes ``_value_for_param(label, ctx)``
      into the cell at column ``4N + 4``.
    - Empty resolver results leave the template cell untouched (so any pre-filled
      defaults in the template survive).
    - Values are coerced via :func:`_coerce_cell_value` so numeric strings render
      as numbers in Excel.
    - When ``revisions`` is a non-empty sequence, the revision title block is
      cleared and the entries are written top-down from
      :data:`TEMPLATE_REVISION_FIRST_ROW` (one revision per row). Pass ``None``
      or ``[]`` to leave the template's built-in title block untouched.

    Projects with more than :data:`TEMPLATE_MAX_LIFTS` lifts still export — only
    the first eight are written and the return value reflects that. The caller is
    responsible for surfacing that truncation to the user.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e

    src = template_path or default_schedule_template_path()
    wb = load_workbook(src)
    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        ws = wb[TEMPLATE_SHEET_NAME]
    else:
        ws = wb.active

    if revisions:
        _write_revision_block(ws, revisions)

    override_fill = PatternFill(
        start_color=OVERRIDE_HIGHLIGHT_HEX,
        end_color=OVERRIDE_HIGHLIGHT_HEX,
        fill_type="solid",
    )
    normalized_overrides: Dict[int, Set[str]] = {}
    if overrides:
        for lift_idx, labels in overrides.items():
            try:
                key = int(lift_idx)
            except (TypeError, ValueError):
                continue
            norm_labels = {
                _normalize_override_label(lbl) for lbl in (labels or ()) if lbl
            }
            norm_labels.discard("")
            if norm_labels:
                normalized_overrides[key] = norm_labels

    lifts_to_write = max(0, min(int(num_lifts), TEMPLATE_MAX_LIFTS))
    last_row = ws.max_row

    for n in range(1, lifts_to_write + 1):
        lift_index = n - 1
        ctx = _build_ctx(user_inputs, lift_index, door_manufacturer=door_manufacturer)
        target_col = _lift_column_index(n)
        lift_overrides = normalized_overrides.get(lift_index, set())

        for r in range(TEMPLATE_PARAM_FIRST_ROW, last_row + 1):
            label_cell = ws.cell(r, 1).value
            if label_cell is None:
                continue
            label = str(label_cell).strip()
            if not label:
                continue
            value = _value_for_param(label, ctx)
            if value in (None, ""):
                continue
            cell = ws.cell(r, target_col)
            # Preserve template-authored Excel formulas (V3.0 introduced cascading
            # formulas like =H82+1.5 / =ROUND(...) / =H53). The template's formula
            # is authoritative for those cells — overwriting them with a static
            # value would break the user's interactive recalc when they edit a
            # dependency cell in Excel. Leave the formula in place.
            existing = cell.value
            if isinstance(existing, str) and existing.startswith("="):
                continue
            cell.value = _coerce_cell_value(value)
            if lift_overrides and _normalize_override_label(label) in lift_overrides:
                cell.fill = override_fill

    wb.save(output_path)
    return lifts_to_write
