"""
Lift Designer → LD import workbook (non-UI).

Export rules come only from ``VT standard configurations_V00.xlsx`` sheet
``VT Standard configs`` column **F** (LD export) when the cell contains an explicit
LD variable path (``Shaft0…``, ``FLL…``, ``L_Projects…``, etc.). Rows marked
``no`` or ``yes, info`` are not exported.

Per **lift index** ``i`` (0 = Lift 1, 1 = Lift 2, …): ``Shaft0`` in paths becomes
``Shaft{i}``. For **Connected load**, **Rated current**, and **Heat dissipation motor**,
``L_Projects.PROJ_USER_*`` is remapped: Lift 1 → 0, 2, 3; Lift 2+ → consecutive triples
starting at ``3*i + 1`` (e.g. Lift 2: 4–6, Lift 3: 7–9, Lift 4: 10–12).

Workbook layout (``SyncWithLD``): row **1** = titles in **A–F** and **Mode declaration** in **G1**;
**G2–G4** hold the three mode lines on the **same** rows as the first content in **A–F** (Shaft0 data may start on row **2**).
Default output file: ``LD data import test.xlsx``.

Use :func:`write_ld_workbook_multi` with the rows for one or more lifts to produce one workbook
(**Shaft *n*** between lifts in that file). Use :func:`write_ld_exports_per_group` to write **one
file per Building System group** (from ``LiftColumnGroups``), each file containing only that group’s lifts.
"""
from __future__ import annotations

import csv
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Mapping, MutableMapping, Optional, Sequence, Tuple, Union

from gui.project_lift_schema import merged_lift_at
from lift_designer_vt_derived import compute_derived

__all__ = [
    "LDExportRow",
    "VTExportRule",
    "load_vt_export_rules",
    "build_ld_rows_from_user_inputs",
    "build_ld_rows_per_lift",
    "matrix_for_ld_workbook",
    "write_ld_workbook",
    "write_ld_workbook_multi",
    "write_ld_exports_per_group",
    "write_ld_csv",
    "export_ld_data_import_test",
    "project_resource_dir",
    "default_vt_workbook_path",
    "default_ld_empty_template_path",
    "default_ld_example_path",
    "remap_varname_for_lift",
]

Number = Union[int, float]


def project_resource_dir() -> str:
    """
    Directory that holds shipped ``.xlsx`` workbooks next to the Python sources.

    When running under PyInstaller (``--onefile``), bundled data files are
    extracted to :attr:`sys._MEIPASS`; this returns that folder so template
    paths resolve correctly.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _default_ld_test_output_path() -> str:
    """Default ``LD data import test.xlsx`` path — writable when frozen (not inside the bundle)."""
    if getattr(sys, "frozen", False):
        base = os.path.join(os.path.expanduser("~"), "Documents")
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "LD data import test.xlsx")


@dataclass(frozen=True)
class LDExportRow:
    varname: str
    value: str = ""
    description: str = ""
    # Source VT row this export row was produced from. Used by
    # :func:`matrix_for_ld_workbook` to anchor bold section-header rows at the correct
    # position. ``0`` means "unknown" and is treated as *before any VT row*.
    vt_row: int = 0


@dataclass
class VTExportRule:
    """One logical rule from the VT sheet (row order preserved)."""

    kind: str  # "static" | "floors_z" | "floors_desc"
    varnames: List[str] = field(default_factory=list)
    param_label: str = ""
    vt_row: int = 0


def default_vt_workbook_path() -> str:
    return os.path.join(project_resource_dir(), "VT standard configurations_V00.xlsx")


def default_ld_empty_template_path() -> str:
    return os.path.join(project_resource_dir(), "LD Export Empty file.xlsx")


def default_ld_example_path() -> str:
    return os.path.join(project_resource_dir(), "LD example data import.xlsx")


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "yes" if v else "no"
    return str(v).strip()


def _norm_param(s: str) -> str:
    return " ".join(str(s).lower().split())


def _is_no_export(cell_val: Any) -> bool:
    if cell_val is None:
        return True
    t = str(cell_val).strip().lower()
    return t in ("", "no", "n", "none")


def _parse_ld_path_lines(cell_val: Any) -> List[str]:
    """Split column F; keep lines that look like LD property paths."""
    if cell_val is None:
        return []
    raw = str(cell_val).strip()
    if not raw or raw in ("-", "?", "yes, info", "yes,info"):
        return []
    out: List[str] = []
    for ln in raw.splitlines():
        ln = ln.strip()
        if not ln or ln == "?":
            continue
        if _is_no_export(ln):
            continue
        if ln.lower() in ("yes, info", "yes,info"):
            continue
        out.append(ln)
    return out


def _looks_like_ld_path(s: str) -> bool:
    """True if ``s`` is an explicit LD path from VT (not ``no`` / ``yes, info`` / ``-``)."""
    s = s.strip()
    if not s:
        return False
    low = s.lower()
    if low in ("no", "n") or low.startswith("no,"):
        return False
    if low in ("yes, info", "yes,info", "-", "?"):
        return False
    return bool(
        re.match(r"^Shaft0(\.|$)", s)
        or s.startswith("FLL.")
        or s.startswith("L_")
        or s.startswith("Shaft.")
    )


def _proj_user_triple_indices(lift_index: int) -> Tuple[int, int, int]:
    """
    ``L_Projects`` suffixes for Connected load, Rated current, Heat dissipation motor.

    Lift 1 (index 0): 0, 2, 3 — Lift k (index i≥1): ``3*i+1``, ``3*i+2``, ``3*i+3``
    (Lift 2: 4–6, Lift 3: 7–9, Lift 4: 10–12, …).
    """
    if lift_index <= 0:
        return (0, 2, 3)
    base = 3 * lift_index + 1
    return (base, base + 1, base + 2)


def remap_varname_for_lift(varname: str, param_label: str, lift_index: int) -> str:
    """
    VT paths use ``Shaft0…``; replace with ``Shaft{n}`` where ``n`` == ``lift_index``.
    Remap ``L_Projects.PROJ_USER_*`` for the three electrical parameters per lift.
    """
    vn = varname.strip()
    norm = _norm_param(param_label)

    if vn.startswith("L_Projects.PROJ_USER_"):
        ic, ir, ih = _proj_user_triple_indices(lift_index)
        if norm == "connected load":
            return f"L_Projects.PROJ_USER_{ic}"
        if norm == "rated current":
            return f"L_Projects.PROJ_USER_{ir}"
        if norm == "heat dissipation motor":
            return f"L_Projects.PROJ_USER_{ih}"
        return vn

    if lift_index == 0:
        return vn
    # Shaft0 → Shaft{lift_index} (VT template is always Shaft0 for lift 1)
    return re.sub(r"\bShaft0\b", f"Shaft{lift_index}", vn)


def load_vt_export_rules(
    path_vt: Optional[str] = None,
) -> List[VTExportRule]:
    """
    Scan ``VT Standard configs`` column A (parameter) and F (LD export).
    """
    path_vt = path_vt or default_vt_workbook_path()

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e

    wb = load_workbook(path_vt, data_only=True)
    if "VT Standard configs" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["VT Standard configs"]
    rules: List[VTExportRule] = []
    seen_z = False
    seen_desc = False

    for r in range(3, ws.max_row + 1):
        param = ws.cell(r, 1).value
        fcell = ws.cell(r, 6).value
        if not param and not fcell:
            continue
        plab = _s(param)
        if not plab:
            continue
        if fcell is None or str(fcell).strip() == "":
            continue
        fs = str(fcell).strip()
        # Only rows whose column F resolves to an explicit DTV path are exported.
        # ``no`` / ``-`` / ``yes`` / ``yes, info`` (and any other free-text marker
        # without an LD variable) are dropped entirely: they must never produce a
        # row in the LD export workbook, even as a description-only placeholder.
        if fs.lower() in ("yes, info", "yes,info"):
            continue
        if _is_no_export(fs):
            continue
        if fs == "-":
            continue
        lines = _parse_ld_path_lines(fs)
        if not lines:
            continue
        paths = [ln.strip() for ln in lines if _looks_like_ld_path(ln.strip())]
        if not paths:
            continue
        if any(re.match(r"^FLL\.Level\d+\.Z_POT$", x) for x in paths):
            if not seen_z:
                rules.append(VTExportRule("floors_z", [], plab, r))
                seen_z = True
            continue
        if any(re.match(r"^FLL\.Level\d+\.DESC$", x) for x in paths):
            if not seen_desc:
                rules.append(VTExportRule("floors_desc", [], plab, r))
                seen_desc = True
            continue
        rules.append(VTExportRule("static", paths, plab, r))

    wb.close()
    return rules


@dataclass
class _ExportCtx:
    user_inputs: Mapping[str, Any]
    lift_index: int
    lift: Dict[str, Any]
    forces: Dict[str, Any]
    drive: Dict[str, Any]
    compliance: Dict[str, Any]
    emergency: Dict[str, Any]
    cost: Dict[str, Any]
    # VT-formula derived values keyed by normalized parameter label (lowercase, whitespace-
    # collapsed). Used as a fallback when the UI field is empty — see :func:`_value_for_param`.
    derived: Dict[str, str] = field(default_factory=dict)


def _lift(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    return merged_lift_at(ui, i) if isinstance(ui, dict) else {}


def _forces(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    rows = ui.get("Forces") or []
    if 0 <= i < len(rows) and isinstance(rows[i], dict):
        return rows[i]
    return {}


def _drive(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    rows = ui.get("LiftDrive") or []
    if 0 <= i < len(rows) and isinstance(rows[i], dict):
        return rows[i]
    return {}


def _compliance(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    rows = ui.get("Compliance") or []
    if 0 <= i < len(rows) and isinstance(rows[i], dict):
        return rows[i]
    return {}


def _emergency(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    rows = ui.get("Emergency") or []
    if 0 <= i < len(rows) and isinstance(rows[i], dict):
        return rows[i]
    return {}


def _cost(ui: Mapping[str, Any], i: int) -> Dict[str, Any]:
    rows = ui.get("Cost") or []
    if 0 <= i < len(rows) and isinstance(rows[i], dict):
        return rows[i]
    return {}


def _floors_list(ui: Mapping[str, Any], i: int) -> List[MutableMapping[str, Any]]:
    floors = ui.get("Floors") or []
    if i >= len(floors) or not isinstance(floors[i], dict):
        return []
    block = floors[i]
    key = f"Lift {i + 1}"
    raw = block.get(key) if key in block else next(iter(block.values()), [])
    return list(raw) if isinstance(raw, list) else []


def _get_lift_field(lift: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = lift.get(k)
        if v is not None and str(v).strip() != "":
            return _s(v)
    return ""


# Applicable codes page: checkbox row labels — export short code names only when checked.
_LD_COMPLIANCE_CHECKBOX_LABELS: Tuple[str, ...] = (
    "EN81-28 emergency call",
    "EN81-70 Accessibility",
    "DIN EN17210 / 18040-1 Accessibility",
    "EN81-72 Firefighter elevator",
    "EN81-73 Fire emergency return",
    "EN81-58 Fire rated landing doors",
)

# Full UI label → standard/code token for LD export (no descriptions or "=value" suffixes).
_LD_COMPLIANCE_CHECKBOX_LABEL_TO_CODE: Dict[str, str] = {
    "EN81-28 emergency call": "EN81-28",
    "EN81-70 Accessibility": "EN81-70",
    "DIN EN17210 / 18040-1 Accessibility": "DIN EN17210 / 18040-1",
    "EN81-72 Firefighter elevator": "EN81-72",
    "EN81-73 Fire emergency return": "EN81-73",
    "EN81-58 Fire rated landing doors": "EN81-58",
}


def _compliance_checkbox_selected(v: Any) -> bool:
    if v is True:
        return True
    if isinstance(v, str) and v.strip().lower() in ("yes", "true", "1"):
        return True
    return False


def _compliance_summary(ui: Mapping[str, Any], i: int) -> str:
    """
    Text for ``L_StandardTab.STD_DESC`` / “Applied codes”: short standard/code names only
    (e.g. ``EN81-70``, ``DIN EN17210 / 18040-1``, ``EN81-76``, ``BREEAM``) — no UI suffixes
    or ``field=value`` text.
    """
    comp = ui.get("Compliance") or []
    if i >= len(comp) or not isinstance(comp[i], dict):
        return ""
    d = comp[i]
    parts: List[str] = []
    seen: set[str] = set()

    def add(token: str) -> None:
        t = token.strip()
        if t and t not in seen:
            seen.add(t)
            parts.append(t)

    for label in _LD_COMPLIANCE_CHECKBOX_LABELS:
        if label not in d:
            continue
        if _compliance_checkbox_selected(d[label]):
            code = _LD_COMPLIANCE_CHECKBOX_LABEL_TO_CODE.get(label, label)
            add(code)

    v = d.get("EN81-71 Vandalism category")
    if v is not None and str(v).strip() not in ("", "0"):
        add("EN81-71")

    en81_76 = False
    v = d.get("EN81-76 Emergency Evacuation type")
    if isinstance(v, str) and v.strip() and v.strip().lower() != "no":
        en81_76 = True
    v = d.get("EN81-76 Evacuation functions")
    if v is not None and str(v).strip():
        en81_76 = True
    if en81_76:
        add("EN81-76")

    v = d.get("EN81-77 Seismic category")
    if v is not None and str(v).strip() not in ("", "0"):
        add("EN81-77")

    v = d.get("Green building certification compliance")
    if v is None:
        pass
    elif isinstance(v, dict):
        for scheme, sel in v.items():
            if _compliance_checkbox_selected(sel):
                add(_s(scheme))
    elif isinstance(v, (list, tuple)):
        for item in v:
            if item is not None and str(item).strip():
                add(_s(item))
    elif str(v).strip():
        add(_s(v))

    v = d.get("EN81-58 Fire rating class")
    if isinstance(v, bool):
        if v:
            add("EN81-58")
    elif isinstance(v, str) and v.strip():
        add("EN81-58")

    return "; ".join(parts)


def _emergency_summary(em: Dict[str, Any]) -> str:
    parts: List[str] = []
    for k, v in em.items():
        if isinstance(v, str) and v.strip():
            parts.append(f"{k}={v}")
    return "; ".join(parts)


# --- Parameter label (VT column A, normalized) → value resolver ---

def _val_system_name(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "System Type") or _s(ctx.user_inputs.get("FileName", ""))


def _val_configuration_short(ctx: _ExportCtx) -> str:
    return _s(ctx.user_inputs.get("FileName", ""))


def _val_counterweight(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Counterweight location")


def _val_load(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Load capacity", "Load capacity (kg)")


def _val_persons(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Permissible number of persons", "Permissible number of persons (Pers.)")


def _val_speed(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Speed", "Speed (m/s)")


def _val_num_floors_fll(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Number of floors", "Number of floors (Stck.)")


def _val_stops(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Stops", "Stops (Stck.)")


def _val_cw(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Cabin width", "Cabin width (mm)")


def _val_cd(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Cabin depth", "Cabin depth (mm)")


def _val_height(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Structural cabin height", "Structural cabin height (mm)")


def _val_door_open_b(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Door structural opening width", "Door structural opening width (mm)")


def _val_head(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Shaft head suggested", "Shaft head suggested (mm)")


def _val_pit(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Shaft pit suggested", "Shaft pit suggested (mm)")


def _val_connected(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.drive, "Connected load", "Connected load (kVA)")


def _val_rated_i(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.drive, "Rated current", "Rated current (A)")


def _val_heat(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.drive, "Heat dissipation motor", "Heat dissipation motor (kJ)")


def _val_clad(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Cladding thickness each wall", "Cladding thickness each wall (mm)")


def _val_force_f12(ctx: _ExportCtx) -> str:
    return _get_lift_field(
        ctx.forces,
        "Force F1, F2 elevator rail segment",
        "Force F1, F2 elevator rail segment (kN)",
    )


def _val_force_f3(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force F3, each buffer", "Force F3, each buffer (kN)")


def _val_force_f4(ctx: _ExportCtx) -> str:
    return _get_lift_field(
        ctx.forces,
        "Force F4, per counterweight rail segment",
        "Force F4, per counterweight rail segment (kN)",
    )


def _val_force_f5(ctx: _ExportCtx) -> str:
    return _get_lift_field(
        ctx.forces,
        "Force F5, per counterweight buffer",
        "Force F5, per counterweight buffer (kN)",
    )


def _val_force_f6(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force F6, static shaft door", "Force F6, static shaft door (kN)")


def _val_force_f7(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force F7, static counterweight", "Force F7, static counterweight (kN)")


def _val_force_f8(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force F8, static cabin", "Force F8, static cabin (kN)")


def _val_force_fx_car(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force Fx, cabin rail", "Force Fx, cabin rail (kN)")


def _val_force_fy_car(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force Fy, cabin rail", "Force Fy, cabin rail (kN)")


def _val_force_fx_cwt(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force Fx, counterweight rail", "Force Fx, counterweight rail (kN)")


def _val_force_fy_cwt(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.forces, "Force Fy, counterweight rail", "Force Fy, counterweight rail (kN)")


def _val_std_desc(ctx: _ExportCtx) -> str:
    return _compliance_summary(ctx.user_inputs, ctx.lift_index)


def _val_lop(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "LOP type and location", "LOP type and locaion")


def _val_lip(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "LIP type and location")


def _val_door_h(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Door height", "Door height (mm)")


def _val_door_type(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "door type")


def _val_shaft_width(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Shaft width suggested", "Shaft width suggested (mm)")


def _val_shaft_depth(ctx: _ExportCtx) -> str:
    return _get_lift_field(ctx.lift, "Shaft depth suggested", "Shaft depth suggested (mm)")


PARAM_RESOLVERS: Dict[str, Callable[[_ExportCtx], str]] = {
    "configuration short name": _val_configuration_short,
    "system name": _val_system_name,
    "load capacity": _val_load,
    "permissible number of persons": _val_persons,
    "speed": _val_speed,
    "number of floors": _val_num_floors_fll,
    "stops": _val_stops,
    "cabin width (clear)": _val_cw,
    "cabin depth (clear)": _val_cd,
    "structural cabin height": _val_height,
    "door structural opening width": _val_door_open_b,
    "shaft head suggested": _val_head,
    "shaft pit suggested": _val_pit,
    "connected load": _val_connected,
    "rated current": _val_rated_i,
    "heat dissipation motor": _val_heat,
    "energy recovery": lambda c: _get_lift_field(c.drive, "Energy recovery", "Energy recovery (y/n)"),
    "diversity factor": lambda c: _get_lift_field(c.drive, "Diversity factor", "Diversity factor (-)"),
    "energy consumption": lambda c: _get_lift_field(c.drive, "Energy consumption", "Energy consumption (kWh)"),
    "force f1, f2 elevator rail segment": _val_force_f12,
    "force f3, each buffer": _val_force_f3,
    "force f4, per counterweight rail segment": _val_force_f4,
    "force f5, per counterweight buffer": _val_force_f5,
    "force f6, static shaft door": _val_force_f6,
    "force f7, static counterweight": _val_force_f7,
    "force f8, static cabin": _val_force_f8,
    "force fx, cabin rail": _val_force_fx_car,
    "force fy, cabint rail": _val_force_fy_car,
    "force fy, cabin rail": _val_force_fy_car,
    "force fx, counterweight rail": _val_force_fx_cwt,
    "force fy, counterweight rail": _val_force_fy_cwt,
    "applied codes": _val_std_desc,
    "cladding thickness each wall": _val_clad,
    "lop type and location": _val_lop,
    "lip type and location": _val_lip,
    "door width": lambda c: _get_lift_field(c.lift, "Door width", "Door width (mm)"),
    "door height": _val_door_h,
    "door type": _val_door_type,
    "shaft width suggested": _val_shaft_width,
    "shaft depth suggested": _val_shaft_depth,
    "counterweight location": _val_counterweight,
}


def _strip_qualifier(label: str) -> str:
    """
    Drop trailing parenthetical qualifiers and unit suffixes and collapse whitespace
    around ``/`` for loose matching.

    - ``"Cabin width (clear)"`` → ``"cabin width"``
    - ``"Door width (mm)"``     → ``"door width"``
    - ``"Accessible rooms / cwt safety"`` → ``"accessible rooms/cwt safety"``
      so it matches the UI key ``"Accessible rooms/cwt safety"``.

    Repeated parentheses are stripped until none remain so labels like
    ``"Acceleration (m/s²) (new)"`` still normalize cleanly.
    """
    s = str(label).strip()
    while True:
        new = re.sub(r"\s*\([^)]*\)\s*$", "", s)
        if new == s:
            break
        s = new
    s = re.sub(r"\s*/\s*", "/", s)
    return _norm_param(s)


def _lookup_value_in_ctx_dicts(param_key: str, ctx: _ExportCtx) -> str:
    """
    Scan every UI-backed dict on ``ctx`` for a key that matches ``param_key`` under
    normal or loose normalization. Returns the first non-empty match, or ``""``.

    The loose form tolerates the common VT / UI mismatch where VT labels carry a
    parenthetical qualifier (e.g. ``"Cabin width (clear)"``) while the UI stores the
    value under the plain name (``"Cabin width"``). Exact-normalized matches are
    preferred over loose ones so a UI that does write ``"Cabin width (clear)"`` still
    wins over one that writes ``"Cabin width"``.
    """
    loose_target = _strip_qualifier(param_key)
    dicts = (ctx.lift, ctx.forces, ctx.drive, ctx.compliance, ctx.emergency, ctx.cost)

    # First pass: exact normalized match.
    for d in dicts:
        if not isinstance(d, Mapping):
            continue
        for dk, dv in d.items():
            if not isinstance(dk, str):
                continue
            if _norm_param(dk) == param_key:
                v = _s(dv)
                if v:
                    return v
    # Second pass: qualifier-stripped match (both sides).
    for d in dicts:
        if not isinstance(d, Mapping):
            continue
        for dk, dv in d.items():
            if not isinstance(dk, str):
                continue
            if _strip_qualifier(dk) == loose_target:
                v = _s(dv)
                if v:
                    return v
    return ""


def _value_for_param(param_label: str, ctx: _ExportCtx) -> str:
    """
    Resolve a VT parameter label to an LD export value.

    Order of precedence (UI always wins):

    1. UI-backed resolver in :data:`PARAM_RESOLVERS`.
    2. Exact / loose-qualifier match across every UI-backed context dict
       (``ctx.lift``, ``ctx.forces``, ``ctx.drive``, …). This catches values the UI
       writes under labels that differ from the VT column-A label only by a trailing
       parenthetical qualifier — e.g. UI ``"Cabin width"`` vs VT ``"Cabin width (clear)"``.
    3. VT-formula fallback in ``ctx.derived`` — only used when the UI has nothing to
       contribute, so manually entered values are never overridden by a computed default.
    """
    key = _norm_param(param_label)

    fn = PARAM_RESOLVERS.get(key)
    if fn:
        val = fn(ctx)
        if val:
            return val

    # Special: L_StandardTab.STD_DESC description "Applied codes"
    if "applied codes" in key:
        std = _val_std_desc(ctx)
        if std:
            return std

    ui_val = _lookup_value_in_ctx_dicts(key, ctx)
    if ui_val:
        return ui_val

    derived = ctx.derived.get(key, "")
    if derived:
        return derived
    # Derived also exposes an alias under the qualifier-stripped label (see
    # :mod:`lift_designer_vt_derived`). Try that as a last resort so e.g. a VT row
    # labelled plainly "Cabin width" still picks up the profile's computed width.
    derived_loose = ctx.derived.get(_strip_qualifier(key), "")
    if derived_loose:
        return derived_loose

    return ""


def _coerce_cell_value(val: str) -> Union[str, Number]:
    if val is None or val == "":
        return ""
    s = str(val).strip()
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except ValueError:
            return s
    if re.fullmatch(r"-?\d+[\.,]\d+", s):
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return s
    return val


def _floor_elevation_mm(floor: Mapping[str, Any]) -> int:
    """
    Read the floor's absolute elevation (metres) and return it in millimetres.

    Accepts the current UI key ``"Elevation (m)"`` and, for backward compatibility with
    projects saved before the rename, falls back to the legacy ``"Height (m)"`` key.
    Non-numeric / missing entries return ``0`` mm.
    """
    raw = floor.get("Elevation (m)")
    if raw is None or str(raw).strip() == "":
        raw = floor.get("Height (m)", 0)
    try:
        return int(round(float(str(raw).replace(",", ".")) * 1000.0))
    except (ValueError, TypeError, OverflowError):
        return 0


def _rows_for_floors_z(ctx: _ExportCtx, vt_row: int = 0) -> List[LDExportRow]:
    """
    Emit ``FLL.Level{i}.Z_POT`` rows. The value is each floor's **absolute elevation**
    (UI field *Elevation (m)*) converted to millimetres — one direct mapping per row, no
    cumulative summing. The description carries the floor name (same text that the
    matching ``.DESC`` row uses for its value).
    """
    out: List[LDExportRow] = []
    floors = _floors_list(ctx.user_inputs, ctx.lift_index)
    for idx, floor in enumerate(floors):
        z_mm = _floor_elevation_mm(floor)
        fname = _s(floor.get("Floor Name", floor.get("Floor", "")))
        label = fname if fname else f"Level {idx}"
        out.append(
            LDExportRow(
                f"FLL.Level{idx}.Z_POT",
                str(z_mm),
                label,
                vt_row=vt_row,
            )
        )
    return out


def _rows_for_floors_desc(ctx: _ExportCtx, vt_row: int = 0) -> List[LDExportRow]:
    """
    Emit ``FLL.Level{i}.DESC`` rows. The value is the UI *Floor Name*; the description
    falls back to the numeric *Floor* if the name is blank so the LD workbook always
    has a readable label in column F.
    """
    out: List[LDExportRow] = []
    floors = _floors_list(ctx.user_inputs, ctx.lift_index)
    for idx, floor in enumerate(floors):
        fn = _s(floor.get("Floor Name", ""))
        fl = _s(floor.get("Floor", ""))
        out.append(LDExportRow(f"FLL.Level{idx}.DESC", fn, fl or fn, vt_row=vt_row))
    return out


def build_ld_rows_from_user_inputs(
    user_inputs: Mapping[str, Any],
    lift_index: int = 0,
    vt_path: Optional[str] = None,
    door_manufacturer: Optional[str] = None,
) -> List[LDExportRow]:
    """
    Build export rows using ``VT standard configurations`` LD export column.

    ``door_manufacturer`` selects the VT R277 manufacturer cell (drives door RID,
    door depth and door/wall clearance lookups). When omitted, the value is taken
    from ``user_inputs["DoorManufacturer"]`` (or the project default).
    """
    ctx = _ExportCtx(
        user_inputs=user_inputs,
        lift_index=lift_index,
        lift=_lift(user_inputs, lift_index),
        forces=_forces(user_inputs, lift_index),
        drive=_drive(user_inputs, lift_index),
        compliance=_compliance(user_inputs, lift_index),
        emergency=_emergency(user_inputs, lift_index),
        cost=_cost(user_inputs, lift_index),
        derived=compute_derived(user_inputs, lift_index, door_manufacturer=door_manufacturer),
    )
    rules = load_vt_export_rules(vt_path)
    rows: List[LDExportRow] = []

    for rule in rules:
        if rule.kind == "floors_z":
            rows.extend(_rows_for_floors_z(ctx, vt_row=rule.vt_row))
            continue
        if rule.kind == "floors_desc":
            rows.extend(_rows_for_floors_desc(ctx, vt_row=rule.vt_row))
            continue

        val = _value_for_param(rule.param_label, ctx)

        for vn in rule.varnames:
            tv = remap_varname_for_lift(vn, rule.param_label, ctx.lift_index)
            rows.append(LDExportRow(tv, val, rule.param_label, vt_row=rule.vt_row))

    return rows


def _section_row_f(title: str) -> List[Any]:
    """Empty row with section title in column F only (bold applied when writing xlsx)."""
    return [None, None, None, None, None, title, None]


# Bold section headers emitted into column F of the LD workbook. Each entry is
# ``(anchor_vt_row, title)`` — the header is inserted just before the first LD row whose
# source VT row is ``>= anchor_vt_row``. Rows are matched in list order, so entries must
# be sorted by ``anchor_vt_row``.
#
# The anchor row is usually the VT row of the header itself (column A), except where a
# header must appear *between* two consecutive data rows. Example: the VT sheet has
# "Lift Designer parameters" at R231 but "Applied codes" (a data row) at R232 — to get
# the order *Applicable codes → Applied codes (STD_DESC) → Lift Designer parameters →
# COP panel*, the "Lift Designer parameters" anchor is R233 (the first row after
# STD_DESC) and "Applicable codes" anchors at R232 (right before STD_DESC).
#
# Parent/sub-section combinations (e.g. "Shaft depth" + "1. Open Through") are merged
# into a single title, matching the user-facing grouping the export is expected to show.
_LD_SECTION_HEADERS: Tuple[Tuple[int, str], ...] = (
    (73,  "Electrical & HVAC"),
    (90,  "Mechanical Loading"),
    (232, "Applicable codes"),
    (233, "Lift Designer parameters"),
    (254, "Level name"),
    (265, "Level name"),
    (276, "Entrance Doors"),
    (284, "Car doors"),
    (291, "Lift arrangement tool"),
    (299, "Shaft Width"),
    (311, "Shaft depth 1. Open Through"),
    (326, "Shaft depth 2. Front entrance"),
    (337, "Shaft depth 3. Rear entrance"),
    (355, "Rail Wall thickness"),
    (361, "Door opening height"),
    (367, "Entrance centering Front Door"),
    (372, "Entrance centering Rear door"),
    (380, "Front LOP"),
    (386, "Rear LOP"),
    (391, "LIP"),
    (400, "Front entrance bottom rails"),
    (405, "Front entrance top rails"),
    (410, "Rear entrance rails"),
)


def _ld_workbook_preamble() -> List[List[Any]]:
    """Row 1 only: **A–F** titles and **G1** = *Mode declaration* (see :func:`_apply_ld_mode_legend_to_rows_g234`)."""
    c, d = "DTV_MODE", "SYNC_MODE"
    return [
        ["DTV_VARNAME", "DTV_VALUE", c, d, None, None, "Mode declaration"],
    ]


_LD_MODE_LEGEND_G234: Tuple[str, str, str] = (
    "1: Read property from Liftdesigner",
    "2: Write back to Liftdesigner",
    "3: Read from LD & write to LD",
)


def _apply_ld_mode_legend_to_rows_g234(matrix: List[List[Any]]) -> None:
    """Set **G** on Excel rows 2–4 (0-based indices 1–3) to the three mode lines; may overlap A–F on those rows."""
    for i in range(3):
        r = 1 + i
        if r >= len(matrix):
            break
        row = matrix[r]
        while len(row) < 7:
            row.append(None)
        row[6] = _LD_MODE_LEGEND_G234[i]


def matrix_for_ld_workbook(
    rows: Sequence[LDExportRow],
    *,
    with_mode_legend: bool = True,
) -> Tuple[List[List[Any]], List[int]]:
    """
    Build a 7-column grid like the reference ``LD example data import.xlsx``:

    - Row **1**: titles in **A–F**, **G1** = *Mode declaration*; **G2–G4** = mode lines on rows **2–4** (with A–F on those rows).
    - Section breaks: empty row with **bold** title in column **F** only.
    - Parameter rows: A–D values, description in **F**, **G** empty.

    Section headers come from :data:`_LD_SECTION_HEADERS` and are anchored to VT rows
    carried on each :class:`LDExportRow`. A header is inserted just before the first
    data row whose ``vt_row`` is ``>= header.anchor_vt_row``. Rows with ``vt_row == 0``
    (legacy / synthetic) are treated as *after every anchor* so they don't trigger
    pending headers on their own.

    ``with_mode_legend`` controls whether the three mode-description lines are written
    into **G2–G4** of this sub-matrix. Multi-lift assembly in
    :func:`matrix_for_ld_workbook_multi` disables it per-lift and stamps the legend
    exactly once on the final combined matrix, so the text stays in rows 1–4 and does
    not leak onto the first data rows of subsequent shaft blocks.

    Returns ``(matrix, section_header_row_indices)`` (1-based row numbers for bold F cells).
    """
    matrix: List[List[Any]] = list(_ld_workbook_preamble())
    section_rows_1based: List[int] = []
    pending = list(_LD_SECTION_HEADERS)  # consumed in order
    seen_titles: set = set()

    def _flush_headers_up_to(target_row: int) -> None:
        while pending and pending[0][0] <= target_row:
            _, title = pending.pop(0)
            # The VT sheet lists "Level name" twice (Z_POT and DESC groups). Keep both
            # occurrences (they anchor to different VT rows) but drop accidental
            # re-adds if the table is ever extended with real duplicates.
            key = (title, target_row)
            if key in seen_titles:
                continue
            seen_titles.add(key)
            matrix.append(_section_row_f(title))
            section_rows_1based.append(len(matrix))

    for r in rows:
        if not str(r.varname).strip() and not str(r.value).strip():
            continue

        if r.vt_row:
            _flush_headers_up_to(r.vt_row)

        matrix.append(
            [
                r.varname,
                _coerce_cell_value(r.value) if isinstance(r.value, str) else r.value,
                0,
                2,
                None,
                r.description or None,
                None,
            ]
        )

    if with_mode_legend:
        _apply_ld_mode_legend_to_rows_g234(matrix)
    return matrix, section_rows_1based


def build_ld_rows_per_lift(
    user_inputs: Mapping[str, Any],
    num_lifts: int,
    vt_path: Optional[str] = None,
    door_manufacturer: Optional[str] = None,
) -> List[List[LDExportRow]]:
    """Return ``num_lifts`` row lists: index ``i`` is ``build_ld_rows_from_user_inputs(..., lift_index=i)``."""
    n = max(0, int(num_lifts))
    return [
        build_ld_rows_from_user_inputs(
            user_inputs,
            lift_index=i,
            vt_path=vt_path,
            door_manufacturer=door_manufacturer,
        )
        for i in range(n)
    ]


def matrix_for_ld_workbook_multi(
    rows_by_lift: Sequence[Sequence[LDExportRow]],
) -> Tuple[List[List[Any]], List[int]]:
    """
    Single sheet: preamble once, then each lift's block. Before lift index ``i`` (``i >= 1``),
    insert a bold **Shaft *i*** row in column **F** (same style as other section titles).

    Each lift uses a fresh copy of the internal section headers (Electrical & HVAC, …).
    The mode-declaration legend is applied **once** at the end (to the global rows 2–4)
    so the three description lines never appear on the first data rows of shafts 2+.
    """
    matrix: List[List[Any]] = list(_ld_workbook_preamble())
    section_row_nums: List[int] = []
    for i, rows in enumerate(rows_by_lift):
        if i > 0:
            matrix.append(_section_row_f(f"Shaft {i}"))
            section_row_nums.append(len(matrix))
        sub_m, sub_sec = matrix_for_ld_workbook(list(rows), with_mode_legend=False)
        body = sub_m[1:]
        offset = len(matrix)
        matrix.extend(body)
        for s in sub_sec:
            if s > 1:
                section_row_nums.append(offset + (s - 1))
    _apply_ld_mode_legend_to_rows_g234(matrix)
    return matrix, section_row_nums


def _safe_ld_group_filename_segment(name: str, index_1based: int) -> str:
    """Sanitize group title for use in a Windows filename segment."""
    s = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", (name or "").strip())
    s = re.sub(r"\s+", "_", s).strip("_") or f"group{index_1based}"
    return s[:80]


def write_ld_exports_per_group(
    base_save_path: str,
    rows_by_lift: Sequence[Sequence[LDExportRow]],
    lift_groups_raw: Optional[Sequence[Mapping[str, Any]]],
    *,
    sheet_title: str = "SyncWithLD",
) -> List[str]:
    """
    Write **one** ``SyncWithLD`` workbook per Building System group.

    ``base_save_path`` is the path chosen in the save dialog. With a single group, the workbook
    is written exactly there. With multiple groups, files are
    ``{stem}_{sanitized_group_title}.xlsx`` in the same directory (one suffix from the group name).

    Row lists must align with ``LiftColumnGroups`` counts (normalized via
    :func:`gui.project_lift_schema.parse_lift_column_groups`).
    Returns all paths written.
    """
    from gui.project_lift_schema import merge_consecutive_lift_groups_same_name, parse_lift_column_groups

    n = len(rows_by_lift)
    groups = parse_lift_column_groups(list(lift_groups_raw) if lift_groups_raw is not None else [], n)
    groups = merge_consecutive_lift_groups_same_name(groups)

    pairs: List[Tuple[Mapping[str, Any], List[Sequence[LDExportRow]]]] = []
    idx = 0
    for g in groups:
        cnt = int(g.get("count", 0))
        if cnt <= 0:
            continue
        chunk = list(rows_by_lift[idx : idx + cnt])
        idx += cnt
        pairs.append((g, chunk))

    if not pairs:
        return []

    if len(pairs) == 1:
        write_ld_workbook_multi(base_save_path, pairs[0][1], sheet_title=sheet_title)
        return [base_save_path]

    parent = os.path.dirname(base_save_path)
    stem = os.path.splitext(os.path.basename(base_save_path))[0]
    written: List[str] = []
    used_lower: set[str] = set()
    for gi, (g, chunk) in enumerate(pairs):
        safe = _safe_ld_group_filename_segment(str(g.get("name", "")), gi + 1)
        fname = f"{stem}_{safe}.xlsx"
        out_path = os.path.join(parent, fname)
        key = out_path.lower()
        if key in used_lower:
            fname = f"{stem}_{safe}_{gi + 1}.xlsx"
            out_path = os.path.join(parent, fname)
            key = out_path.lower()
        used_lower.add(key)
        write_ld_workbook_multi(out_path, chunk, sheet_title=sheet_title)
        written.append(out_path)
    return written


def _matrix_for_workbook(rows: Sequence[LDExportRow]) -> List[List[Any]]:
    """Backward-compatible wrapper; section styling is applied in :func:`write_ld_workbook`."""
    m, _ = matrix_for_ld_workbook(rows)
    return m


def _rows_to_matrix_csv(rows: Sequence[LDExportRow]) -> List[List[str]]:
    """Legacy CSV (no preamble legend in A)."""
    out: List[List[str]] = [
        ["DTV_VARNAME", "DTV_VALUE", "DTV_MODE", "SYNC_MODE", "", "", ""],
    ]
    for r in rows:
        if not str(r.varname).strip() and not str(r.value).strip():
            continue
        out.append([r.varname, _s(r.value), "0", "2", "", r.description, ""])
    return out


def write_ld_csv(path: str, rows: Sequence[LDExportRow]) -> None:
    matrix = _rows_to_matrix_csv(rows)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for line in matrix:
            w.writerow(line)


def _write_ld_matrix_to_path(
    path: str,
    matrix: List[List[Any]],
    section_row_nums: Sequence[int],
    sheet_title: str = "SyncWithLD",
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError("pip install openpyxl") from e
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    font_header = Font(bold=True)
    font_section = Font(bold=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    def _alignment_for_column(c_idx: int) -> Alignment:
        # A, F, G: left; B, C, D: center; E: left (unused spacer)
        if c_idx in (1, 5, 6, 7):
            return align_left
        return align_center

    for r_idx, line in enumerate(matrix, start=1):
        for c_idx, cell in enumerate(line, start=1):
            if c_idx > 7:
                continue
            cl = ws.cell(row=r_idx, column=c_idx, value=cell)
            cl.alignment = _alignment_for_column(c_idx)
            if r_idx == 1 and cell is not None:
                cl.font = font_header
            if r_idx in section_row_nums and c_idx == 6 and cell is not None:
                cl.font = font_section

    for col in range(1, 8):
        letter = get_column_letter(col)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2.5, 12), 85)

    wb.save(path)


def write_ld_workbook_multi(
    path: str,
    rows_by_lift: Sequence[Sequence[LDExportRow]],
    sheet_title: str = "SyncWithLD",
) -> None:
    """Write one ``SyncWithLD`` sheet: preamble once, then each lift block (**Shaft *n*** between lifts)."""
    matrix, section_row_nums = matrix_for_ld_workbook_multi(rows_by_lift)
    _write_ld_matrix_to_path(path, matrix, section_row_nums, sheet_title=sheet_title)


def write_ld_workbook(
    path: str,
    rows: Sequence[LDExportRow],
    sheet_title: str = "SyncWithLD",
) -> None:
    write_ld_workbook_multi(path, [rows], sheet_title=sheet_title)


def export_ld_data_import_test(
    user_inputs: Mapping[str, Any],
    lift_index: int = 0,
    output_path: Optional[str] = None,
    vt_path: Optional[str] = None,
) -> str:
    """
    Write ``LD data import test.xlsx`` next to this module (or ``output_path``).
    Returns the path written.
    """
    out = output_path or _default_ld_test_output_path()
    rows = build_ld_rows_from_user_inputs(
        user_inputs,
        lift_index=lift_index,
        vt_path=vt_path,
    )
    write_ld_workbook(out, rows)
    return out


if __name__ == "__main__":
    import argparse
    import sys

    p = argparse.ArgumentParser(description="Export project data to LD data import Excel (SyncWithLD).")
    p.add_argument(
        "project_json",
        nargs="?",
        default=None,
        help="Path to a LiftDesign project .json file. If omitted, a minimal demo dict is used.",
    )
    p.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output .xlsx path (default: LD data import test.xlsx next to this script).",
    )
    p.add_argument(
        "-l",
        "--lift",
        type=int,
        default=0,
        help="Lift index (0-based). Ignored if --all-lifts. Default: 0.",
    )
    p.add_argument(
        "--all-lifts",
        action="store_true",
        help="Write LD workbook(s): one file per lift group (LiftColumnGroups), or one file if there is a single group.",
    )
    args = p.parse_args()

    if args.project_json:
        try:
            from gui.project_json import load_project_json
        except ImportError:
            print("Import error: run from the project root so the gui package is found.", file=sys.stderr)
            raise SystemExit(1) from None
        data = load_project_json(args.project_json)
    else:
        if args.all_lifts:
            data = {
                "FileName": "demo",
                "BuildingSystems": [{}, {}],
                "GeneralSpecification": [{}, {}],
                "LayoutInformation": [{}, {}],
                "LiftDrive": [{}, {}],
                "Forces": [{}, {}],
                "Compliance": [{}, {}],
                "Floors": [{"Lift 1": []}, {"Lift 2": []}],
            }
        else:
            data = {
                "FileName": "demo",
                "BuildingSystems": [{}],
                "GeneralSpecification": [{}],
                "LayoutInformation": [{}],
                "LiftDrive": [{}],
                "Forces": [{}],
                "Compliance": [{}],
                "Floors": [{"Lift 1": []}],
            }

    out = args.output or _default_ld_test_output_path()
    if args.all_lifts:
        try:
            from gui.project_lift_schema import KEY_LIFT_COLUMN_GROUPS
        except ImportError:
            print("Import error: run from the project root so the gui package is found.", file=sys.stderr)
            raise SystemExit(1) from None
        n = len(data.get("BuildingSystems") or [])
        rows_by_lift = build_ld_rows_per_lift(data, n)
        paths = write_ld_exports_per_group(out, rows_by_lift, data.get(KEY_LIFT_COLUMN_GROUPS))
        for p in paths:
            print(p)
        path = paths[-1] if paths else out
    else:
        path = export_ld_data_import_test(
            data,
            lift_index=args.lift,
            output_path=args.output,
        )
    print(path)
